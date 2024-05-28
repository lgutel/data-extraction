import openpyxl
import os
from SQL import connexionsql
import pandas as pd
def fichierleplusrecent(chemin):
    fichiers_excel = [fichier for fichier in os.listdir(chemin) if fichier.lower().endswith(".xlsx")]

    if not fichiers_excel:
        print("Aucun fichier Excel trouvé dans le répertoire spécifié.")
    else:
        fichiers_excel.sort(key=lambda fichier: os.path.getmtime(os.path.join(chemin, fichier)), reverse=True)
        return fichiers_excel[0]

def nomfeuille(chemin, fichier_recent):
    excel_file_path = os.path.join(chemin, fichier_recent)
    wb = openpyxl.load_workbook(excel_file_path)

    # Obtenez la liste des noms de feuilles
    sheet_names = wb.sheetnames
    print("Voici une liste des noms de feuille :")
    for sheet_name in sheet_names:
        print(f"- {sheet_name}")

def lire_feuille(chemin, fichier_recent, feuille):
    excel_file_path = os.path.join(chemin, fichier_recent)
    wb = openpyxl.load_workbook(excel_file_path)

    try:
        feuille_selectionnee = wb[feuille]
        donnees = []
        for row in feuille_selectionnee.iter_rows():
            ligne = [cell.value for cell in row]
            donnees.append(ligne)

        # Créer un DataFrame à partir des données
        df = pd.DataFrame(donnees)

        # Afficher le DataFrame dans le terminal
        print(df)
    except Exception as e:
        print(f"Erreur lors de la lecture de la feuille : {e}")

def main():
    while True:
        g = input("Que voulez-vous faire ? (extraction/connexion) ")
        if g.lower() == "extraction":
            chemin = input(r"Entrez le chemin de votre document (attention à la casse) : ")
            fichier_recent = fichierleplusrecent(chemin)
            a = input("Est-ce qu'il y a plusieurs feuilles dans ce document ? ")
            if a.lower() == 'oui':
                nomfeuille(chemin, fichier_recent)
                feuille = input("Entrez le nom de la feuille à extraire : ")
                lire_feuille(chemin, fichier_recent, feuille)
            else:
                feuille = input("Entrez le nom de la feuille à extraire : ")
                lire_feuille(chemin, fichier_recent, feuille)
        elif g.lower() == "connexion":
            connexionsql()
        else:
            break

if __name__ == '__main__':
    main()
